from flask import Flask, render_template, request, redirect, send_file
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.utils import get_column_letter
import psycopg2
import os

app = Flask(__name__)
MATA_UANG = ['USD', 'IDR', 'KHR']

# Format angka gaya Indonesia
def format_angka(value):
    try:
        return "{:,.2f}".format(float(value)).replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return value

app.jinja_env.filters['format_angka'] = format_angka

# Koneksi ke PostgreSQL
DB_PARAMS = {
    'dbname': 'keuangan',
    'user': 'admin',
    'password': os.environ.get('DB_PASSWORD'),
    'host': 'dpg-d1cgko6uk2gs73an2t50-a',
    'port': 5432
}

def get_connection():
    return psycopg2.connect(**DB_PARAMS)

def simpan_transaksi(tipe, deskripsi, mata_uang, jumlah):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO transaksi (tanggal, tipe, deskripsi, mata_uang, jumlah)
        VALUES (%s, %s, %s, %s, %s)
    """, (datetime.now().strftime('%Y-%m-%d'), tipe, deskripsi, mata_uang, jumlah))
    conn.commit()
    cur.close()
    conn.close()

def get_transaksi_hari_ini():
    tanggal = datetime.now().strftime('%Y-%m-%d')
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, tanggal, tipe, deskripsi, mata_uang, jumlah
        FROM transaksi WHERE tanggal = %s
    """, (tanggal,))
    rows = [
        {'id': r[0], 'Tanggal': r[1], 'Tipe': r[2], 'Deskripsi': r[3], 'Mata Uang': r[4], 'Jumlah': r[5]}
        for r in cur.fetchall()
    ]
    cur.close()
    conn.close()
    return rows

def get_semua_transaksi():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, tanggal, tipe, deskripsi, mata_uang, jumlah FROM transaksi")
    rows = [
        {'id': r[0], 'Tanggal': r[1], 'Tipe': r[2], 'Deskripsi': r[3], 'Mata Uang': r[4], 'Jumlah': r[5]}
        for r in cur.fetchall()
    ]
    cur.close()
    conn.close()
    return rows

def ringkasan_hari_ini():
    tanggal = datetime.now().strftime('%Y-%m-%d')
    pemasukan = defaultdict(float)
    pengeluaran = defaultdict(float)

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT tipe, mata_uang, jumlah FROM transaksi WHERE tanggal = %s", (tanggal,))
    for tipe, mata_uang, jumlah in cur.fetchall():
        if tipe.lower() == 'pemasukan':
            pemasukan[mata_uang] += jumlah
        else:
            pengeluaran[mata_uang] += jumlah
    cur.close()
    conn.close()

    omset = {mu: pemasukan[mu] - pengeluaran[mu] for mu in MATA_UANG}
    return pemasukan, pengeluaran, omset

def saldo_per_mata_uang():
    saldo = defaultdict(float)
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT tipe, mata_uang, jumlah FROM transaksi")
    for tipe, mata_uang, jumlah in cur.fetchall():
        if tipe.lower() == 'pemasukan':
            saldo[mata_uang] += jumlah
        else:
            saldo[mata_uang] -= jumlah
    cur.close()
    conn.close()
    return saldo

def saldo_utama():
    kurs = {'USD': 15500, 'KHR': 3.8, 'IDR': 1}
    saldo = saldo_per_mata_uang()
    return sum(saldo[m] * kurs[m] for m in MATA_UANG)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        tipe = request.form['tipe']
        deskripsi = request.form['deskripsi']
        mata_uang = request.form['mata_uang']
        jumlah = request.form['jumlah'].replace(',', '.')
        try:
            jumlah = float(jumlah)
            simpan_transaksi(tipe, deskripsi, mata_uang, jumlah)
        except:
            pass
        return redirect('/')

    pemasukan, pengeluaran, omset = ringkasan_hari_ini()
    saldo = saldo_utama()
    return render_template('index.html',
                           pemasukan=pemasukan,
                           pengeluaran=pengeluaran,
                           omset=omset,
                           saldo=saldo,
                           saldo_per_mata_uang=saldo_per_mata_uang(),
                           mata_uang=MATA_UANG)

@app.route('/riwayat')
def riwayat():
    data = get_transaksi_hari_ini()
    return render_template('riwayat.html', data=data)

@app.route('/admin')
def admin():
    data = get_semua_transaksi()
    return render_template('admin.html', data=data)

@app.route('/download')
def download_excel():
    rows = get_transaksi_hari_ini()
    if not rows:
        return "Tidak ada data hari ini.", 404

    nama_file = "riwayat_hari_ini.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Riwayat Hari Ini"

    headers = ['Tanggal', 'Tipe', 'Deskripsi', 'Mata Uang', 'Jumlah']
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    wb.save(nama_file)
    return send_file(nama_file, as_attachment=True)

@app.route('/hapus', methods=['POST'])
def hapus():
    id_transaksi = request.form.get('id')
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM transaksi WHERE id = %s", (id_transaksi,))
    conn.commit()
    cur.close()
    conn.close()
    return redirect(request.referrer)

# Tambahkan route ini ke app.py
@app.route('/edit/<int:id>', methods=['GET', 'POST'])
def edit(id):
    conn = get_connection()
    cur = conn.cursor()

    if request.method == 'POST':
        tipe = request.form['tipe']
        deskripsi = request.form['deskripsi']
        mata_uang = request.form['mata_uang']
        jumlah = float(request.form['jumlah'].replace(',', '.'))
        cur.execute("""
            UPDATE transaksi
            SET tipe = %s, deskripsi = %s, mata_uang = %s, jumlah = %s
            WHERE id = %s
        """, (tipe, deskripsi, mata_uang, jumlah, id))
        conn.commit()
        cur.close()
        conn.close()
        return redirect('/')

    cur.execute("SELECT tipe, deskripsi, mata_uang, jumlah FROM transaksi WHERE id = %s", (id,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    return render_template("edit.html", id=id, data=row, mata_uang=MATA_UANG)


def buat_table_transaksi():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS transaksi (
            id SERIAL PRIMARY KEY,
            tanggal DATE,
            tipe VARCHAR(20),
            deskripsi TEXT,
            mata_uang VARCHAR(5),
            jumlah FLOAT
        )
    """)
    conn.commit()
    cur.close()
    conn.close()

if __name__ == '__main__':
    buat_table_transaksi()
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
